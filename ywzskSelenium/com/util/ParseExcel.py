# coding = utf-8
import openpyxl
from openpyxl.styles import Border,Side,Font
import time
class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)
        self.RGBDict = {'red':'FFFF3030','green':'FF008B00'}
    def loadWorkBook(self,excelPathAndName):
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception ,e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook
    def getSheetByName(self,sheetName):
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception,e:
            raise e
    def getSheetByIndex(self,sheetIndex):
        try:
            sheetname = self.workbook.get_sheet_by_names()[sheetIndex]
        except Exception,e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self,sheet):
        return sheet.max_row
    def getColsNumber(self,sheet):
        return sheet.max_column
    def getStartRowNumber(self,sheet):
        return  sheet.min_row
    def getStartColsNumber(self,sheet):
        return sheet.min_column
    def getRow(self,sheet,rowNo):
        try:
            return sheet.rows[rowNo-1]
        except Exception,e:
            raise e
    def getColnum(self,sheet,colNo):
        try:
            return sheet.colnums[colNo-1]
        except Exception,e:
            raise e
    def getCellOfValue(self,sheet,coordinate=None,rowNo = None,colsNo = None):
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate).value
            except Exception,e:
                raise e
        elif coordinate is None and rowNo and colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column=colsNo).value
            except Exception,e:
                raise e

        else:
            raise Exception("Insufficint Coordinates of cell!")

    def getCellOfObject(self,sheet,coordinate=None,rowNo = None,colsNo = None):
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate)
            except Exception,e:
                raise e
        elif coordinate ==None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo)
            except Exception,e:
                raise e
        else:
            raise Exception("Insufficint Coordinates of cell!")

    def writeCell(self,sheet,content,coordinate = None,rowNo = None,colsNo = None,style = None):
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate = coordinate).font = Font(color= self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value = content
                if style:
                    sheet.cell(row=rowNo,column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        else:
            raise Exception("Insufficint Coordinates of cell!")

    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo = None,colsNo = None):
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        elif coordinate==None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
            else:
                raise Exception("Insufficint Coordinates of cell!")




